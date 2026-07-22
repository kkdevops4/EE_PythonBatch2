import os
import pandas as pd
from openpyxl import load_workbook

class DataLoader:
    def __init__(self, file_path):
        self.file_path = file_path

    def show_excel_sheets(self):
        workbook = load_workbook(self.file_path, read_only=True)
        print("\nAvailable Excel Sheets:")
        for sheet_name in workbook.sheetnames:
            print("-", sheet_name)
        workbook.close()

    def load_data(self):
        if not os.path.exists(self.file_path):
            print("Excel file not found:", self.file_path)
            return None
        try:
            self.show_excel_sheets()
            data = pd.read_excel(self.file_path,sheet_name="Driver Monitoring Data",engine="openpyxl")
            required = ["Driver_ID", "Driver_Name", "Shift", "Time_Minutes",
                "Travel_Distance_km", "Eye_Closure_%", "Blink_Rate",
                "Head_Pitch_Angle", "Yawning_Count"]
            missing = [column for column in required if column not in data.columns]
            if missing:
                print("Missing required columns:", missing)
                return None

            numeric = ["Time_Minutes", "Travel_Distance_km", "Eye_Closure_%",
                "Blink_Rate", "Head_Pitch_Angle", "Yawning_Count"]
            for column in numeric:
                data[column] = pd.to_numeric(data[column], errors="coerce")

            data[numeric] = data[numeric].fillna(0)
            print("\nDataset Loaded Successfully.")
            print("Total Records:", len(data))
            return data

        except Exception as error:
            print("Error while loading data:", error)
            return None